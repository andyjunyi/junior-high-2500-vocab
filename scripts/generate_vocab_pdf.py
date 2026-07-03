#!/usr/bin/env python3
"""Generate a well-designed PDF from the 初中英語2500詞講義."""

import openpyxl
import json
import html

# Load data
wb = openpyxl.load_workbook('/srv/projects/junior-high-2500-vocab/data/初中英語2500詞_自學講義.xlsx')

ws = wb['單字表']
words = []
for row in ws.iter_rows(min_row=2, values_only=True):
    if row[1] is None:
        continue
    words.append({
        'id': row[0], 'word': str(row[1]), 'phonetic': str(row[2]) if row[2] else '',
        'pos': str(row[3]) if row[3] else '', 'meaning': str(row[4]) if row[4] else '',
        'example': str(row[5]) if row[5] else '', 'example_zh': str(row[6]) if row[6] else '',
        'letter': str(row[7]) if row[7] else '', 'note': str(row[8]) if row[8] else ''
    })

# Organize by letter
from collections import defaultdict
by_letter = defaultdict(list)
for w in words:
    by_letter[w['letter']].append(w)

letters = sorted(by_letter.keys())

# Learning guide
ws2 = wb['學習指引']
guide_items = []
for row in ws2.iter_rows(values_only=True):
    if row[0]:
        guide_items.append(str(row[0]))

# Stats
ws3 = wb['統計總覽']
stats = []
for row in ws3.iter_rows(values_only=True):
    stats.append([str(c) if c else '' for c in row])

# ── HTML generation ──

def esc(text):
    if text is None:
        return ''
    return html.escape(str(text))

def build_word_card(w):
    """Single word entry as a compact card."""
    phonetic = f' <span class="phonetic">{esc(w["phonetic"])}</span>' if w['phonetic'] else ''
    pos = f' <span class="pos">{esc(w["pos"])}</span>' if w['pos'] else ''
    meaning = f' <span class="meaning">{esc(w["meaning"])}</span>' if w['meaning'] else ''
    
    parts = []
    parts.append(f'''<div class="word-card">
    <div class="word-head">
      <span class="word-num">{w['id']}</span>
      <span class="word-text">{esc(w['word'])}</span>{phonetic}{pos}
    </div>
    <div class="word-meaning">{meaning}</div>''')
    
    if w['example']:
        parts.append(f'''    <div class="word-example">
      <span class="example-en">📝 {esc(w['example'])}</span>
      <span class="example-zh">{esc(w['example_zh'])}</span>
    </div>''')
    
    if w['note'] and w['note'] != 'None':
        parts.append(f'    <div class="word-note">💡 {esc(w["note"])}</div>')
    
    parts.append('  </div>')
    return '\n'.join(parts)

# Build letter sections
letter_sections = []
for letter in letters:
    group = by_letter[letter]
    cards = '\n'.join(build_word_card(w) for w in group)
    letter_sections.append(f'''
  <section class="letter-section">
    <h2 class="letter-header">{esc(letter)} <span class="letter-count">({len(group)} 字)</span></h2>
    <div class="word-grid">
{cards}
    </div>
  </section>''')

# Build guide section
guide_html_parts = []
for item in guide_items:
    if item.startswith('📚'):
        guide_html_parts.append(f'<h2 class="guide-title-main">{esc(item)}</h2>')
    elif item.startswith('📌'):
        guide_html_parts.append(f'<h3 class="guide-title">{esc(item)}</h3>')
    elif item.startswith('以下'):
        guide_html_parts.append(f'<p class="guide-subtitle">{esc(item)}</p>')
    elif item.strip().startswith('*') or item.strip().startswith('-'):
        guide_html_parts.append(f'<li class="guide-li">{esc(item.strip().lstrip("* -"))}</li>')
    elif item.strip():
        guide_html_parts.append(f'<p class="guide-p">{esc(item)}</p>')
    else:
        guide_html_parts.append('<br>')
guide_html = '\n'.join(guide_html_parts)

# Build stats
stats_html = '<table class="stats-table">'
for row in stats:
    cells = ''.join(f'<td>{esc(c)}</td>' for c in row)
    stats_html += f'<tr>{cells}</tr>'
stats_html += '</table>'

# Total counts
total_words = len(words)
words_with_example = sum(1 for w in words if w['example'])
words_with_note = sum(1 for w in words if w['note'] and w['note'] != 'None')

html_content = f'''<!DOCTYPE html>
<html lang="zh-TW">
<head>
<meta charset="UTF-8">
<title>初中英語 2500 詞自學講義</title>
<style>
  @page {{
    size: A4;
    margin: 0;
    
    @bottom-center {{
      content: "— " counter(page) " —";
      font-family: "Noto Serif CJK TC", serif;
      font-size: 8pt;
      color: #8b9dc3;
    }}
  }}
  
  @page :first {{
    @bottom-center {{
      content: none;
    }}
  }}
  
  * {{ margin: 0; padding: 0; box-sizing: border-box; }}
  
  body {{
    font-family: "Noto Serif CJK TC", "Noto Sans CJK TC", serif;
    font-size: 9pt;
    line-height: 1.5;
    color: #2c3e50;
    counter-reset: vocab-counter;
  }}

  /* ── Cover ── */
  .cover {{
    page: cover;
    height: 297mm;
    width: 210mm;
    display: flex;
    flex-direction: column;
    justify-content: center;
    align-items: center;
    background: linear-gradient(135deg, #1a1a2e 0%, #16213e 40%, #0f3460 100%);
    color: #ffffff;
    text-align: center;
    position: relative;
    overflow: hidden;
  }}
  
  .cover::before {{
    content: "";
    position: absolute;
    top: -50%;
    left: -50%;
    width: 200%;
    height: 200%;
    background: radial-gradient(circle at 30% 70%, rgba(233, 69, 96, 0.08) 0%, transparent 60%),
                radial-gradient(circle at 70% 30%, rgba(72, 149, 239, 0.08) 0%, transparent 60%);
  }}
  
  .cover-content {{
    position: relative;
    z-index: 1;
  }}
  
  .cover-emoji {{
    font-size: 48pt;
    margin-bottom: 20px;
    display: block;
  }}
  
  .cover h1 {{
    font-size: 28pt;
    font-weight: 700;
    letter-spacing: 0.05em;
    margin-bottom: 10px;
    text-shadow: 0 2px 10px rgba(0,0,0,0.3);
  }}
  
  .cover .subtitle {{
    font-size: 12pt;
    color: #a8c8f0;
    font-weight: 300;
    margin-bottom: 30px;
  }}
  
  .cover .stats-badge {{
    display: inline-flex;
    gap: 24px;
    font-size: 10pt;
    color: #7ba5d9;
  }}
  
  .cover .stats-badge span {{
    background: rgba(255,255,255,0.06);
    padding: 6px 16px;
    border-radius: 20px;
    border: 1px solid rgba(255,255,255,0.1);
  }}

  /* ── TOC ── */
  .toc-page {{
    padding: 25mm 20mm;
    page-break-after: always;
  }}
  
  .toc-title {{
    font-size: 16pt;
    font-weight: 700;
    color: #1a1a2e;
    margin-bottom: 15px;
    padding-bottom: 8px;
    border-bottom: 2px solid #0f3460;
  }}
  
  .toc-grid {{
    display: grid;
    grid-template-columns: repeat(5, 1fr);
    gap: 6px 12px;
  }}
  
  .toc-item {{
    font-size: 9pt;
    padding: 4px 8px;
    background: #f0f4fa;
    border-radius: 4px;
    text-align: center;
    color: #2c3e50;
  }}
  
  .toc-item strong {{
    color: #0f3460;
  }}

  /* ── Letter section ── */
  .letter-section {{
    padding: 8mm 12mm 4mm 12mm;
    page-break-before: always;
  }}
  
  .letter-header {{
    font-size: 16pt;
    font-weight: 700;
    color: #1a1a2e;
    margin-bottom: 8px;
    padding-bottom: 4px;
    border-bottom: 2px solid #0f3460;
    display: flex;
    align-items: baseline;
    gap: 8px;
  }}
  
  .letter-count {{
    font-size: 9pt;
    color: #7f8c8d;
    font-weight: 400;
  }}
  
  .word-grid {{
    display: grid;
    grid-template-columns: 1fr 1fr;
    gap: 4px 10px;
  }}

  /* ── Word card ── */
  .word-card {{
    background: #fafbfd;
    border: 1px solid #e8ecf1;
    border-radius: 5px;
    padding: 5px 7px;
    font-size: 8.5pt;
    break-inside: avoid;
    transition: background 0.2s;
  }}
  
  .word-head {{
    display: flex;
    align-items: baseline;
    gap: 4px;
    flex-wrap: wrap;
    margin-bottom: 1px;
  }}
  
  .word-num {{
    font-size: 6.5pt;
    color: #bdc3c7;
    min-width: 16px;
  }}
  
  .word-text {{
    font-weight: 700;
    color: #0f3460;
    font-size: 9.5pt;
  }}
  
  .phonetic {{
    color: #7f8c8d;
    font-size: 7.5pt;
  }}
  
  .pos {{
    color: #e74c3c;
    font-size: 7pt;
    font-weight: 600;
    background: #fdedec;
    padding: 0 4px;
    border-radius: 2px;
  }}
  
  .word-meaning {{
    color: #2c3e50;
    font-weight: 600;
    font-size: 8.5pt;
  }}
  
  .word-example {{
    margin-top: 2px;
    padding: 3px 5px;
    background: #f7f9fc;
    border-left: 2px solid #3498db;
    border-radius: 0 3px 3px 0;
  }}
  
  .example-en {{
    color: #2c3e50;
    font-style: italic;
    display: block;
    font-size: 8pt;
  }}
  
  .example-zh {{
    color: #95a5a6;
    font-size: 7.5pt;
    display: block;
    margin-top: 1px;
  }}
  
  .word-note {{
    margin-top: 2px;
    font-size: 7pt;
    color: #e67e22;
    font-style: italic;
  }}

  /* ── Guide section ── */
  .guide-section {{
    padding: 15mm 20mm;
    page-break-before: always;
  }}
  
  .guide-title-main {{
    font-size: 16pt;
    font-weight: 700;
    color: #1a1a2e;
    margin-bottom: 4px;
  }}
  
  .guide-subtitle {{
    font-size: 9pt;
    color: #7f8c8d;
    margin-bottom: 15px;
  }}
  
  .guide-title {{
    font-size: 11pt;
    font-weight: 700;
    color: #0f3460;
    margin-top: 12px;
    margin-bottom: 4px;
    padding: 6px 10px;
    background: #edf2f9;
    border-radius: 4px;
    border-left: 4px solid #3498db;
  }}
  
  .guide-p {{
    font-size: 9pt;
    margin-bottom: 6px;
    line-height: 1.6;
  }}
  
  .guide-li {{
    font-size: 8.5pt;
    margin-left: 16px;
    margin-bottom: 3px;
    list-style: disc;
    line-height: 1.5;
  }}

  /* ── Stats section ── */
  .stats-section {{
    padding: 15mm 20mm;
    page-break-before: always;
  }}
  
  .stats-title {{
    font-size: 16pt;
    font-weight: 700;
    color: #1a1a2e;
    margin-bottom: 12px;
  }}
  
  .stats-table {{
    width: 100%;
    border-collapse: collapse;
    font-size: 8.5pt;
  }}
  
  .stats-table td {{
    padding: 4px 8px;
    border-bottom: 1px solid #e8ecf1;
  }}
  
  .stats-table tr:first-child td {{
    font-weight: 700;
    background: #f0f4fa;
    color: #1a1a2e;
  }}

  .final-note {{
    margin-top: 20px;
    padding: 12px;
    background: linear-gradient(135deg, #f0f4fa, #e8f0fe);
    border-radius: 8px;
    font-size: 9pt;
    color: #2c3e50;
    text-align: center;
    border: 1px solid #d0dff5;
  }}
</style>
</head>
<body>

<!-- ═══ COVER ═══ -->
<div class="cover">
  <div class="cover-content">
    <span class="cover-emoji">📚</span>
    <h1>初中英語 2500 詞</h1>
    <p class="subtitle">自學講義 · 繁體中文版</p>
    <div class="stats-badge">
      <span>📝 {total_words} 個單字</span>
      <span>📖 {words_with_example} 句例句</span>
      <span>🏷️ {len(letters)} 個字母章節</span>
    </div>
  </div>
</div>

<!-- ═══ TABLE OF CONTENTS ═══ -->
<div class="toc-page">
  <h2 class="toc-title">📑 目錄索引</h2>
  <div class="toc-grid">
'''

for letter in letters:
    count = len(by_letter[letter])
    html_content += f'    <div class="toc-item"><strong>{esc(letter)}</strong> — {count} 字</div>\n'

html_content += '''  </div>
  <p style="margin-top: 20px; font-size: 9pt; color: #7f8c8d; text-align: center;">
    翻至對應字母章節即可查閱該字母所有單字
  </p>
</div>
'''

# Vocabulary sections
html_content += '\n'.join(letter_sections)

# Guide section
html_content += f'''
<!-- ═══ LEARNING GUIDE ═══ -->
<div class="guide-section">
{guide_html}
</div>
'''

# Stats section
html_content += f'''
<!-- ═══ STATISTICS ═══ -->
<div class="stats-section">
  <h2 class="stats-title">📊 統計總覽</h2>
  {stats_html}
  <div class="final-note">
    📚 本講義依據中國初中英語 2500 詞大綱整理，共收錄 <strong>{total_words}</strong> 個單字，<br>
    每字附 KK 音標、詞性、中文釋義、例句與中譯。適合自學與課堂教學使用。
  </div>
</div>
'''

html_content += '''
</body>
</html>
'''

# Write HTML
html_path = '/tmp/vocab_2500.html'
with open(html_path, 'w', encoding='utf-8') as f:
    f.write(html_content)

print(f"HTML written: {html_path}")
print(f"Words: {total_words}, Letters: {len(letters)}")
